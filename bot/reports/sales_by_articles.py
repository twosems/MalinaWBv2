from aiogram import Router, F
from aiogram.types import CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram_calendar import SimpleCalendar, SimpleCalendarCallback
import math
from storage.users import get_user_api_key
from bot.services.wildberries_api import get_sales_report_for_period, get_stocks
from bot.utils.pagination import build_pagination_keyboard
from bot.utils.calendar import ( remove_builtin_calendar_buttons )
import logging
from storage.articles import get_all_articles  # добавь, если ещё нет
from storage.articles import get_in_stock_articles
from keyboards.keyboards import sales_price_type_keyboard
from storage.users import get_user_price_type, set_user_price_type
from bot.keyboards.keyboards import price_type_human

router = Router()
PAGE_SIZE_ARTICLES = 10
PAGE_SIZE_REPORT = 5

# === Уникальные состояния для артикула ===
class ArticlePeriodCalendarFSM(StatesGroup):
    waiting_for_start = State()
    waiting_for_end = State()

class ArticleDayCalendarFSM(StatesGroup):
    choosing_day = State()

def get_simple_calendar():
    return SimpleCalendar(locale="ru")

# --- Форматирование Telegram-отчёта ---
   # ← вот это обязательно!

def format_sales_report(art, date_from, date_to, stat, price_type_name, page=1, page_size=PAGE_SIZE_REPORT):
    wh_list = list(stat.items())
    total_pages = max(1, math.ceil(len(wh_list) / page_size))
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    end = start + page_size
    current_items = wh_list[start:end]

    grand_qty = sum(sum(s['qty'] for s in sales) for sales in stat.values())
    grand_sum = sum(sum(s['sum'] for s in sales) for sales in stat.values())

    lines = [
        f"📦 Артикул: <code>{art}</code>",
        f"💶 Цена: {price_type_name}",
        f"📅 Период: {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}",
        f"📄 Страница: {page}/{total_pages}",
        ""
    ]
    for wh, sales in current_items:
        qty_sum = sum(s['qty'] for s in sales)
        sum_all = sum(s['sum'] for s in sales)
        price_view = sales[0]['price'] if sales else 0
        lines.append(f"🏬 <b>{wh}</b>")
        lines.append(f"• Кол-во: {qty_sum}")
        lines.append(f"• Цена: {price_view:,.2f}".replace(",", " ").replace(".", ","))
        lines.append(f"• Сумма: {sum_all:,.2f}".replace(",", " ").replace(".", ","))
        lines.append("")

    if page == total_pages:
        lines.append(f"🧾 <b>ИТОГО</b>: {grand_qty} шт / {grand_sum:,.2f} ₽".replace(",", " ").replace(".", ","))

    return "\n".join(lines), total_pages

def format_csv_number(val):
    try:
        return f"{float(val):.2f}".replace(".", ",")
    except Exception:
        return str(val)

# --- Основные обработчики ---
from storage.users import get_user_article_mode

@router.callback_query(F.data == "sales_by_articles")
async def open_sales_by_articles_menu(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    article_mode = await get_user_article_mode(user_id)
    # Получаем нужный список артикулов
    if article_mode == "all":
        arts = sorted(await get_all_articles(user_id))
        label = "Все артикулы продавца"
    else:
        arts = sorted(await get_in_stock_articles(user_id))
        label = "Артикулы с остатком"
    logging.info(f"[ARTICLES] open_sales_by_articles_menu: взято {len(arts)} артикулов ({article_mode}) для user_id={user_id}")

    total = len(arts)
    pages = max(1, (total + PAGE_SIZE_ARTICLES - 1) // PAGE_SIZE_ARTICLES)
    page = 1  # всегда первая страница

    items = arts[(page-1)*PAGE_SIZE_ARTICLES : page*PAGE_SIZE_ARTICLES]
    kb = [
        [InlineKeyboardButton(text=art, callback_data=f"article_period:{art}")]
        for art in items
    ]
    nav_kb = build_pagination_keyboard(
        total=total,
        page=page,
        per_page=PAGE_SIZE_ARTICLES,
        prefix="articles_page:",
        back_callback="main_sales"
    )
    kb.extend(nav_kb.inline_keyboard)

    await callback.message.answer(
        f"{label} ({page}/{pages}):",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )
    await state.clear()

@router.callback_query(F.data.startswith("articles_page:"))
async def articles_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split(":")[1])
    user_id = callback.from_user.id
    article_mode = await get_user_article_mode(user_id)
    if article_mode == "all":
        arts = sorted(await get_all_articles(user_id))
        label = "Все артикулы продавца"
    else:
        arts = sorted(await get_in_stock_articles(user_id))
        label = "Артикулы с остатком"

    total = len(arts)
    pages = max(1, (total + PAGE_SIZE_ARTICLES - 1) // PAGE_SIZE_ARTICLES)
    page = max(1, min(page, pages))
    items = arts[(page-1)*PAGE_SIZE_ARTICLES : page*PAGE_SIZE_ARTICLES]

    kb = [
        [InlineKeyboardButton(text=art, callback_data=f"article_period:{art}")]
        for art in items
    ]
    nav_kb = build_pagination_keyboard(
        total=total,
        page=page,
        per_page=PAGE_SIZE_ARTICLES,
        prefix="articles_page:",
        back_callback="main_sales"
    )
    kb.extend(nav_kb.inline_keyboard)
    await callback.message.edit_text(
        f"{label} ({page}/{pages}):",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )

@router.callback_query(F.data.startswith("article_period:"))
async def article_period_menu(callback: CallbackQuery, state: FSMContext):
    # Сначала пытаемся достать артикул из FSM
    data = await state.get_data()
    art = data.get("article")

    # Если вдруг артикул не найден, попробуем достать из callback.data
    if not art and callback.data and callback.data.startswith("article_period:"):
        art = callback.data.split(":", 1)[1]
        await state.update_data(article=art)

    # Если вообще не нашли артикул — вернуться к общему меню
    if not art:
        await open_sales_by_articles_menu(callback, state)
        return

    kb = [
        [InlineKeyboardButton(text="🗓 За период", callback_data="article_choose_period")],
        [InlineKeyboardButton(text="📅 За день", callback_data="article_choose_day")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="sales_by_articles")]
    ]
    await callback.message.answer(
        f"<b>Артикул:</b> <code>{art}</code>\nВыберите вариант:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb),
        parse_mode="HTML"
    )


# --- Календарь для артикула: период ---
@router.callback_query(F.data == "article_choose_period")
async def article_choose_period(callback: CallbackQuery, state: FSMContext):
    await state.update_data(context_type="article")
    await state.set_state(ArticlePeriodCalendarFSM.waiting_for_start)
    await start_article_period_calendar(callback, state)

@router.callback_query(SimpleCalendarCallback.filter(), ArticlePeriodCalendarFSM.waiting_for_start)
async def article_period_start(callback: CallbackQuery, callback_data: SimpleCalendarCallback, state: FSMContext):
    calendar = get_simple_calendar()
    is_selected, start_date = await calendar.process_selection(callback, callback_data)
    if is_selected:
        await state.update_data(period_start=start_date)
        await state.set_state(ArticlePeriodCalendarFSM.waiting_for_end)
        kb = await calendar.start_calendar()
        kb = remove_builtin_calendar_buttons(kb)
        kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Отмена", callback_data="period_calendar_cancel")])
        await callback.message.edit_text(
            f"✅ <b>Начальная дата:</b> <code>{start_date.strftime('%d.%m.%Y')}</code>\n<b>Теперь выберите конец.</b>",
            reply_markup=kb,
            parse_mode="HTML"
        )

@router.callback_query(SimpleCalendarCallback.filter(), ArticlePeriodCalendarFSM.waiting_for_end)
async def article_period_end(callback: CallbackQuery, callback_data: SimpleCalendarCallback, state: FSMContext):
    calendar = get_simple_calendar()
    is_selected, end_date = await calendar.process_selection(callback, callback_data)
    if is_selected:
        data = await state.get_data()
        start_date = data.get("period_start")
        if end_date < start_date:
            await callback.message.answer("⚠️ Конец должен быть позже начала.")
            return await start_article_period_calendar(callback, state)
        await callback.message.edit_text(
            f"✅ <b>Выбрано:</b> <code>{start_date.strftime('%d.%m.%Y')}</code> — <code>{end_date.strftime('%d.%m.%Y')}</code>\n⏳ Готовим...",
            parse_mode="HTML"
        )
        # --- ОДИН ЗАПРОС к WB, сохраняем в state ---
        await show_sales_article_report(callback, state, start_date, end_date, page=1)
        # Не очищаем state здесь!

@router.callback_query(F.data == "article_choose_day")
async def article_choose_day(callback: CallbackQuery, state: FSMContext):
    await state.update_data(context_type="article")
    await state.set_state(ArticleDayCalendarFSM.choosing_day)
    await start_article_day_calendar(callback, state)

@router.callback_query(SimpleCalendarCallback.filter(), ArticleDayCalendarFSM.choosing_day)
async def article_day_choose(callback: CallbackQuery, callback_data: SimpleCalendarCallback, state: FSMContext):
    calendar = get_simple_calendar()
    is_selected, date = await calendar.process_selection(callback, callback_data)
    if is_selected:
        await callback.message.edit_text(f"✅ <b>Дата:</b> <code>{date.strftime('%d.%m.%Y')}</code>\n⏳ Готовим...", parse_mode="HTML")
        await show_sales_article_report(callback, state, date, date, page=1)
        # Не очищаем state здесь!

# --- Функции запуска календарей ---
async def start_article_period_calendar(callback, state):
    calendar = get_simple_calendar()
    kb = await calendar.start_calendar()
    kb = remove_builtin_calendar_buttons(kb)
    kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Отмена", callback_data="period_calendar_cancel")])
    await callback.message.edit_text(
        "<b>Выберите начальную дату периода:</b>\nИли воспользуйтесь быстрым выбором 👇",
        reply_markup=kb, parse_mode="HTML"
    )

async def start_article_day_calendar(callback, state):
    calendar = get_simple_calendar()
    kb = await calendar.start_calendar()
    kb = remove_builtin_calendar_buttons(kb)
    today = calendar.date.isoformat() if hasattr(calendar, "date") else ""
    if today:
        kb.inline_keyboard.insert(0, [InlineKeyboardButton(text="🗓 Сегодня", callback_data=f"calendar:{today}")])
    kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Отмена", callback_data="calendar_day_cancel")])
    await callback.message.edit_text("🗓 <b>Выберите дату</b>", reply_markup=kb, parse_mode="HTML")

# --- Основная функция формирования и показа отчёта ---
async def show_sales_article_report(callback, state, date_from, date_to, page: int = 1):
    data = await state.get_data()
    art = data.get("article")
    if page == 1 or not data.get("article_report"):
        user_id = callback.from_user.id
        api_key = await get_user_api_key(user_id)
        sales = await get_sales_report_for_period(api_key, date_from, date_to)
        price_type = await get_user_price_type(user_id)
        price_type_name = price_type_human(price_type)
        # ВАЖНО! Получаем выбранный пользователем тип цены
        price_type = await get_user_price_type(user_id)

        stat = {}
        for item in sales:
            if str(item.get("supplierArticle")) != str(art):
                continue
            wh_name = item.get("warehouseName", "—")
            price_val = float(item.get(price_type, 0))
            qty_val = int(item.get("quantity", 1))
            stat.setdefault(wh_name, [])
            stat[wh_name].append({
                "qty": qty_val,
                "price": price_val,
                "sum": price_val * qty_val,
            })
        await state.update_data(article_report={
            "stat": stat,
            "art": art,
            "date_from": date_from,
            "date_to": date_to,
        })
    else:
        # Берём сохранённые в FSM данные (без нового запроса!)
        stat = data["article_report"]["stat"]
        art = data["article_report"]["art"]
        date_from = data["article_report"]["date_from"]
        date_to = data["article_report"]["date_to"]

    text, total_pages = format_sales_report(
        art, date_from, date_to, stat, price_type_name, page=page, page_size=PAGE_SIZE_REPORT
    )
    kb = build_pagination_keyboard(
        total=len(stat),
        page=page,
        per_page=PAGE_SIZE_REPORT,
        prefix="article_report_page:",
        back_callback="sales_by_articles",
        add_export=True,
        export_callback_data="export_article_csv"
    )
    await callback.message.answer(text, reply_markup=kb, parse_mode="HTML")


# --- ПАГИНАЦИЯ (страницы отчёта) ---
@router.callback_query(F.data.startswith("article_report_page:"))
async def article_report_pagination(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split(":")[1])
    data = await state.get_data()
    report = data.get("article_report")
    if not report or not report.get("stat"):
        await callback.message.answer("Нет данных для отчёта.")
        return

    # Получаем user_id, price_type и его текстовое описание
    user_id = callback.from_user.id
    price_type = await get_user_price_type(user_id)
    price_type_name = price_type_human(price_type)

    text, total_pages = format_sales_report(
        report["art"],
        report["date_from"],
        report["date_to"],
        report["stat"],
        price_type_name,           # ← теперь передаём сюда!
        page=page,
        page_size=PAGE_SIZE_REPORT
    )

    kb = build_pagination_keyboard(
        total=len(report["stat"]),
        page=page,
        per_page=PAGE_SIZE_REPORT,
        prefix="article_report_page:",
        back_callback="sales_by_articles",
        add_export=True,
        export_callback_data="export_article_csv"
    )
    await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML")


# --- ОТМЕНА периода ---
# --- ОТМЕНА выбора периода ---
@router.callback_query(F.data == "period_calendar_cancel", ArticlePeriodCalendarFSM.waiting_for_start)
@router.callback_query(F.data == "period_calendar_cancel", ArticlePeriodCalendarFSM.waiting_for_end)
async def article_period_cancel(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    art = data.get("article")
    if art:
        await article_period_menu(callback, state)  # возвращаем в меню артикула
    else:
        await open_sales_by_articles_menu(callback, state)
    # НЕ нужно await state.clear() здесь, иначе потеряешь артикул

# --- ОТМЕНА дня ---
@router.callback_query(F.data == "calendar_day_cancel", ArticleDayCalendarFSM.choosing_day)
async def article_day_cancel(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    art = data.get("article")
    if art:
        await article_period_menu(callback, state)
    else:
        await open_sales_by_articles_menu(callback, state)
    # И здесь не нужно await state.clear()
# --- Экспорт в CSV ---
@router.callback_query(F.data == "export_article_csv")
async def export_article_xlsx(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    report = data.get("article_report")
    if not report or not report.get("stat"):
        await callback.message.answer("Нет данных для экспорта.")
        return

    user_id = callback.from_user.id
    price_type = await get_user_price_type(user_id)
    from bot.keyboards.keyboards import price_type_human
    price_type_name = price_type_human(price_type)

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Вид цены в отчёте
    ws.append([f"Цена: {price_type_name}"])
    ws.append([])

    # Шапка таблицы
    ws.append(["Артикул", "Период", "Склад", "Кол-во", "Цена", "Сумма"])
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    art = report["art"]
    date_from = report["date_from"].strftime("%d.%m.%Y")
    date_to = report["date_to"].strftime("%d.%m.%Y")
    stat = report["stat"]

    # Данные
    for wh, sales in stat.items():
        for sale in sales:
            ws.append([
                art,
                f"{date_from} — {date_to}",
                wh,
                sale["qty"],
                sale["price"],
                sale["sum"]
            ])

    # Итоги
    grand_qty = sum(sum(s['qty'] for s in sales) for sales in stat.values())
    grand_sum = sum(sum(s['sum'] for s in sales) for sales in stat.values())

    ws.append([])
    ws.append(["", "", "ИТОГО", grand_qty, "", grand_sum])
    for i, cell in enumerate(ws[ws.max_row], 1):
        if cell.value == "ИТОГО":
            cell.font = Font(bold=True)

    # Красиво растянуть колонки
    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max(10, max_length + 2)

    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from aiogram.types.input_file import BufferedInputFile
    file = BufferedInputFile(output.read(), "sales_article_report.xlsx")
    await callback.message.answer_document(file, caption="Отчёт по продажам (XLSX)")
