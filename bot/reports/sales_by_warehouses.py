from aiogram import Router, F
from aiogram.types import CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram_calendar import SimpleCalendar, SimpleCalendarCallback
from datetime import datetime
from storage.warehouses import get_cached_warehouses_dicts
from storage.users import get_user_api_key
from bot.services.wildberries_api import get_sales_report_for_period
from bot.utils.pagination import build_pagination_keyboard
from bot.utils.calendar import remove_builtin_calendar_buttons
from utils.text_utils import normalize_warehouse_name
from bot.services.wildberries_api import get_sales_report_with_eta
import io
from aiogram.types.input_file import BufferedInputFile
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from storage.users import get_user_warehouse_filter
import asyncio
from storage.users import get_user_price_type
from bot.keyboards.keyboards import price_type_human

router = Router()
PAGE_SIZE_WAREHOUSES = 10
PAGE_SIZE_REPORT = 10
PAGE_SIZE = 10

class WarehousePeriodCalendarFSM(StatesGroup):
    waiting_for_start = State()
    waiting_for_end = State()

class WarehouseDayCalendarFSM(StatesGroup):
    choosing_day = State()
class AllWarehousesPeriodCalendarFSM(StatesGroup):
    waiting_for_start = State()
    waiting_for_end = State()

class AllWarehousesDayCalendarFSM(StatesGroup):
    choosing_day = State()
def get_simple_calendar():
    return SimpleCalendar(locale="ru")

# --- Главное меню ---
async def open_sales_by_warehouses_menu(callback: CallbackQuery, state: FSMContext = None):
    kb = [
        [InlineKeyboardButton(text="Выбрать склад", callback_data="sales_wh_select:1")],
        [InlineKeyboardButton(text="Все склады", callback_data="sales_all_wh")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="main_sales")]
    ]
    await callback.message.edit_text(
        "<b>Продажи по складам</b>\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb),
        parse_mode="HTML"
    )
    if state:
        await state.clear()
@router.callback_query(F.data == "main_sales_by_warehouses")
async def main_sales_by_warehouses_menu(callback: CallbackQuery, state: FSMContext):
    await open_sales_by_warehouses_menu(callback, state)

# --- Пагинация по складам ---
@router.callback_query(F.data.startswith("sales_wh_select:"))
async def sales_wh_select(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split(":")[1])
    user_id = callback.from_user.id

    # Получаем фильтр и склады
    warehouse_filter = await get_user_warehouse_filter(user_id)
    warehouses = await get_cached_warehouses_dicts()
    if warehouse_filter == "no_sc":
        warehouses = [w for w in warehouses if not w["name"].startswith("СЦ")]

    total = len(warehouses)
    pages = max(1, (total + PAGE_SIZE_WAREHOUSES - 1) // PAGE_SIZE_WAREHOUSES)
    page = max(1, min(page, pages))
    start = (page - 1) * PAGE_SIZE_WAREHOUSES
    end = start + PAGE_SIZE_WAREHOUSES
    items = warehouses[start:end]
    kb = [
        [InlineKeyboardButton(text=w['name'], callback_data=f"sales_wh_menu:{w['id']}")]
        for w in items
    ]
    nav_kb = build_pagination_keyboard(
        total, page, PAGE_SIZE_WAREHOUSES,
        "sales_wh_select:", "main_sales_by_warehouses"
    )
    kb.extend(nav_kb.inline_keyboard)
    await callback.message.edit_text(
        f"Выберите склад ({page}/{pages}):",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb)
    )


# --- Меню по складу: выбор "день"/"период" ---
@router.callback_query(F.data.startswith("sales_wh_menu:"))
async def sales_wh_menu(callback: CallbackQuery, state: FSMContext):
    warehouse_id = callback.data.split(":")[1]
    await state.update_data(sales_wh_id=warehouse_id, context_type="warehouse")
    kb = [
        [InlineKeyboardButton(text="🗓 За период", callback_data="sales_wh_period")],
        [InlineKeyboardButton(text="📅 За день", callback_data="sales_wh_day")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="sales_wh_select:1")]
    ]
    await callback.message.answer(
        "<b>Выберите период для отчёта:</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb),
        parse_mode="HTML"
    )

# --- Функции запуска календарей (как у артикула) ---
async def start_warehouse_period_calendar(callback, state, data=None):
    calendar = get_simple_calendar()
    kb = await calendar.start_calendar()
    kb = remove_builtin_calendar_buttons(kb)
    kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Отмена", callback_data="period_calendar_cancel")])
    await callback.message.edit_text(
        "<b>Выберите начальную дату периода:</b>\nИли воспользуйтесь быстрым выбором 👇",
        reply_markup=kb, parse_mode="HTML"
    )

async def start_warehouse_day_calendar(callback, state, data=None, text="Выберите дату:"):
    calendar = get_simple_calendar()
    kb = await calendar.start_calendar()
    kb = remove_builtin_calendar_buttons(kb)
    kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Отмена", callback_data="calendar_day_cancel")])
    await callback.message.edit_text(text, reply_markup=kb, parse_mode="HTML")

@router.callback_query(F.data == "sales_wh_period")
async def sales_wh_period(callback: CallbackQuery, state: FSMContext):
    user_data = await state.get_data()
    warehouse_id = user_data.get("sales_wh_id")
    await state.update_data(context_type="warehouse", sales_wh_id=warehouse_id)
    await state.set_state(WarehousePeriodCalendarFSM.waiting_for_start)
    await start_warehouse_period_calendar(callback, state, {"sales_wh_id": warehouse_id})

@router.callback_query(F.data == "sales_wh_day")
async def sales_wh_day(callback: CallbackQuery, state: FSMContext):
    user_data = await state.get_data()
    warehouse_id = user_data.get("sales_wh_id")
    await state.update_data(context_type="warehouse", sales_wh_id=warehouse_id)
    await state.set_state(WarehouseDayCalendarFSM.choosing_day)
    await start_warehouse_day_calendar(callback, state, {"sales_wh_id": warehouse_id}, text="Выберите день для отчёта по складу:")

# --- Календарь "за период" для склада ---
@router.callback_query(SimpleCalendarCallback.filter(), WarehousePeriodCalendarFSM.waiting_for_start)
async def period_calendar_start_wh(callback: CallbackQuery, callback_data: SimpleCalendarCallback, state: FSMContext):
    calendar = get_simple_calendar()
    is_selected, start_date = await calendar.process_selection(callback, callback_data)
    if is_selected:
        await state.update_data(period_start=start_date)
        await state.set_state(WarehousePeriodCalendarFSM.waiting_for_end)
        kb = await calendar.start_calendar()
        kb = remove_builtin_calendar_buttons(kb)
        kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Отмена", callback_data="period_calendar_cancel")])
        await callback.message.edit_text(
            f"✅ <b>Начальная дата:</b> <code>{start_date.strftime('%d.%m.%Y')}</code>\n<b>Теперь выберите конец.</b>",
            reply_markup=kb, parse_mode="HTML"
        )

@router.callback_query(SimpleCalendarCallback.filter(), WarehousePeriodCalendarFSM.waiting_for_end)
async def period_calendar_end_wh(callback: CallbackQuery, callback_data: SimpleCalendarCallback, state: FSMContext):
    calendar = get_simple_calendar()
    is_selected, end_date = await calendar.process_selection(callback, callback_data)
    if is_selected:
        data = await state.get_data()
        start_date = data.get("period_start")
        if end_date < start_date:
            await callback.message.answer("⚠️ Конец должен быть позже начала.")
            return await start_warehouse_period_calendar(callback, state)

        await show_sales_report(callback, data.get("sales_wh_id"), start_date, end_date, all_warehouses=False, state=state, page=1)

@router.callback_query(SimpleCalendarCallback.filter(), WarehouseDayCalendarFSM.choosing_day)
async def day_selected_wh(callback: CallbackQuery, callback_data: SimpleCalendarCallback, state: FSMContext):
    calendar = get_simple_calendar()
    is_selected, date = await calendar.process_selection(callback, callback_data)
    if is_selected:
        data = await state.get_data()
       # await callback.message.edit_text(
        #    f"✅ <b>Дата:</b> <code>{date.strftime('%d.%m.%Y')}</code>\n⏳ Готовим...", parse_mode="HTML"
        #)
        await show_sales_report(callback, data.get("sales_wh_id"), date, date, all_warehouses=False, state=state, page=1)

# --- Обработчики отмены календаря ---
@router.callback_query(F.data == "period_calendar_cancel", WarehousePeriodCalendarFSM.waiting_for_start)
@router.callback_query(F.data == "period_calendar_cancel", WarehousePeriodCalendarFSM.waiting_for_end)
async def warehouse_period_cancel(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    wh_id = data.get("sales_wh_id")
    if wh_id:
        await sales_wh_menu(callback, state)
    else:
        await open_sales_by_warehouses_menu(callback, state)

@router.callback_query(F.data == "calendar_day_cancel", WarehouseDayCalendarFSM.choosing_day)
async def warehouse_day_cancel(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    wh_id = data.get("sales_wh_id")
    if wh_id:
        await sales_wh_menu(callback, state)
    else:
        await open_sales_by_warehouses_menu(callback, state)

# --- Пагинация по артикулам внутри склада ---
@router.callback_query(F.data.startswith("sales_wh_page:"))
async def sales_wh_page(callback: CallbackQuery, state: FSMContext):
    try:
        page_str = callback.data.split(":")[1]
        page = int(page_str)
    except (IndexError, ValueError):
        await callback.message.answer("Ошибка пагинации: некорректный номер страницы!")
        return
    data = await state.get_data()
    warehouse_id = data.get("warehouse_id")
    date_from = data.get("date_from")
    date_to = data.get("date_to")
    sales = data.get("sales_report", [])
    warehouses = await get_cached_warehouses_dicts()
    wh = next((w for w in warehouses if str(w["id"]) == str(warehouse_id)), None)
    wh_name = wh.get("name", f"ID {warehouse_id}") if wh else f"ID {warehouse_id}"
    wh_name_norm = normalize_warehouse_name(wh_name)

    # Фильтрация и разбиение на страницы
    stat = {}
    for item in sales:
        sale_warehouse = normalize_warehouse_name(item.get("warehouseName", ""))
        if sale_warehouse != wh_name_norm:
            continue
        art = item.get("supplierArticle", "—")
        price = float(item.get("priceWithDisc", 0))
        if art not in stat:
            stat[art] = {"qty": 0, "sum": 0.0, "price": price}
        stat[art]["qty"] += 1
        stat[art]["sum"] += price

    arts = list(stat.items())
    total_pages = max(1, (len(arts) + PAGE_SIZE_REPORT - 1) // PAGE_SIZE_REPORT)
    page = max(1, min(page, total_pages))
    start = (page - 1) * PAGE_SIZE_REPORT
    end = start + PAGE_SIZE_REPORT
    arts_page = arts[start:end]

    text = (
            f"📦 <b>Продажи по складу: {wh_name}</b>\n"
            f"🗓 <b>Период:</b> {date_from} — {date_to}\n\n"
            f"{'Артикул':<14} | {'Кол-во':>6} | {'Цена':>8} | {'Сумма':>10}\n"
            + '-'*42 + '\n'
    )
    total_qty = total_sum = 0
    for art, st in arts_page:
        text += f"{art:<14} | {st['qty']:>6} | {st['price']:>8.2f} | {st['sum']:>10.2f}\n"
        total_qty += st['qty']
        total_sum += st['sum']

    text += (
        f"\n<b>Итого на странице:</b> {total_qty} шт, {total_sum:.2f} ₽"
        f"\nСтраница {page}/{total_pages}"
    )

    kb = build_pagination_keyboard(len(arts), page, PAGE_SIZE_REPORT, "sales_wh_page:", f"sales_wh_menu:{warehouse_id}")
    kb.inline_keyboard.append([InlineKeyboardButton(text="📥 Экспорт в Excel", callback_data="sales_wh_export_xlsx")])
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=kb)
from math import ceil

@router.callback_query(F.data == "sales_all_wh")
async def sales_all_wh(callback: CallbackQuery, state: FSMContext):
    kb = [
        [InlineKeyboardButton(text="🗓 За период", callback_data="sales_all_wh_period")],
        [InlineKeyboardButton(text="📅 За день", callback_data="sales_all_wh_day")],
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="main_sales_by_warehouses")]
    ]
    await callback.message.answer(
        "<b>Продажи по всем складам</b>\nВыберите период:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=kb),
        parse_mode="HTML"
    )
    await state.clear()

@router.callback_query(F.data == "sales_all_wh_period")
async def sales_all_wh_period(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AllWarehousesPeriodCalendarFSM.waiting_for_start)
    calendar = get_simple_calendar()
    kb = await calendar.start_calendar()
    kb = remove_builtin_calendar_buttons(kb)
    kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Отмена", callback_data="all_period_calendar_cancel")])
    await callback.message.edit_text(
        "<b>Выберите начальную дату периода:</b>\nИли воспользуйтесь быстрым выбором 👇",
        reply_markup=kb, parse_mode="HTML"
    )

@router.callback_query(F.data == "sales_all_wh_day")
async def sales_all_wh_day(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AllWarehousesDayCalendarFSM.choosing_day)
    calendar = get_simple_calendar()
    kb = await calendar.start_calendar()
    kb = remove_builtin_calendar_buttons(kb)
    kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Отмена", callback_data="all_calendar_day_cancel")])
    await callback.message.edit_text("Выберите день для отчёта по всем складам:", reply_markup=kb, parse_mode="HTML")

@router.callback_query(SimpleCalendarCallback.filter(), AllWarehousesPeriodCalendarFSM.waiting_for_start)
async def all_period_calendar_start(callback: CallbackQuery, callback_data: SimpleCalendarCallback, state: FSMContext):
    calendar = get_simple_calendar()
    is_selected, start_date = await calendar.process_selection(callback, callback_data)
    if is_selected:
        await state.update_data(all_period_start=start_date)
        await state.set_state(AllWarehousesPeriodCalendarFSM.waiting_for_end)
        kb = await calendar.start_calendar()
        kb = remove_builtin_calendar_buttons(kb)
        kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Отмена", callback_data="all_period_calendar_cancel")])
        await callback.message.edit_text(
            f"✅ <b>Начальная дата:</b> <code>{start_date.strftime('%d.%m.%Y')}</code>\n<b>Теперь выберите конец.</b>",
            reply_markup=kb, parse_mode="HTML"
        )

@router.callback_query(SimpleCalendarCallback.filter(), AllWarehousesPeriodCalendarFSM.waiting_for_end)
async def all_period_calendar_end(callback: CallbackQuery, callback_data: SimpleCalendarCallback, state: FSMContext):
    calendar = get_simple_calendar()
    is_selected, end_date = await calendar.process_selection(callback, callback_data)
    if is_selected:
        data = await state.get_data()
        start_date = data.get("all_period_start")
        if end_date < start_date:
            await callback.message.answer("⚠️ Конец должен быть позже начала.")
            return await sales_all_wh_period(callback, state)
        await show_sales_report_all_warehouses(callback, start_date, end_date, state=state, page=1)

@router.callback_query(SimpleCalendarCallback.filter(), AllWarehousesDayCalendarFSM.choosing_day)
async def all_day_selected(callback: CallbackQuery, callback_data: SimpleCalendarCallback, state: FSMContext):
    calendar = get_simple_calendar()
    is_selected, date = await calendar.process_selection(callback, callback_data)
    if is_selected:
        await show_sales_report_all_warehouses(callback, date, date, state=state, page=1)

# --- Основная функция ETA ---
async def send_report_eta(callback, date_from, date_to):
    if date_from == date_to:
        period_text = f"<b>{date_from.strftime('%d.%m.%Y')}</b>"
    else:
        period_text = f"<b>{date_from.strftime('%d.%m.%Y')}</b> — <b>{date_to.strftime('%d.%m.%Y')}</b>"
    await callback.message.answer(
        f"✅ Формируем отчёт за период {period_text}.\n"
        f"   Пожалуйста, дождитесь завершения выгрузки! ⏳",
        parse_mode="HTML"
    )

from storage.users import get_user_price_type
from bot.keyboards.keyboards import price_type_human

async def show_sales_report(
        callback,
        warehouse_id,
        date_from,
        date_to,
        all_warehouses=False,
        state: FSMContext = None,
        page: int = 1
):
    page_size = 30
    user_id = callback.from_user.id

    # Получаем тип цены пользователя
    price_type = await get_user_price_type(user_id)
    price_type_name = price_type_human(price_type)

    data = await state.get_data()
    if page == 1 or not data.get("sales_report"):
        if date_from == date_to:
            period_text = f"<b>{date_from.strftime('%d.%m.%Y')}</b>"
        else:
            period_text = f"<b>{date_from.strftime('%d.%m.%Y')}</b> — <b>{date_to.strftime('%d.%m.%Y')}</b>"

        progress_message = await callback.message.answer(
            f"✅ Формируем отчёт за период {period_text}.\n"
            f"💶 <b>Цена:</b> {price_type_name}\n"
            f"   Пожалуйста, дождитесь завершения выгрузки! ⏳",
            parse_mode="HTML"
        )

        api_key = await get_user_api_key(user_id)

        while True:
            result = await get_sales_report_with_eta(api_key, date_from, date_to)
            if isinstance(result, dict) and result.get("error") == "ratelimit":
                retry = result["retry"]
                await progress_message.edit_text(
                    f"✅ Формируем отчёт за период {period_text}.\n"
                    f"💶 <b>Цена:</b> {price_type_name}\n"
                    f"   Отчет будет загружен не ранее чем через: <b>{retry} сек.</b> ⏳",
                    parse_mode="HTML"
                )
                await asyncio.sleep(retry)
                continue
            elif isinstance(result, dict) and result.get("error"):
                await progress_message.edit_text("❌ Ошибка при запросе отчёта.", parse_mode="HTML")
                return
            else:
                sales = result
                break

        await state.update_data(
            sales_report=sales,
            date_from=date_from,
            date_to=date_to,
            warehouse_id=warehouse_id
        )
    else:
        sales = data["sales_report"]

    warehouses = await get_cached_warehouses_dicts()
    wh = next((w for w in warehouses if str(w["id"]) == str(warehouse_id)), None)
    wh_name = wh.get("name", f"ID {warehouse_id}") if wh else f"ID {warehouse_id}"
    wh_name_norm = normalize_warehouse_name(wh_name)

    # --- Группируем по артикулам: qty, price, sum
    stat = {}
    for item in sales:
        sale_warehouse = normalize_warehouse_name(item.get("warehouseName", ""))
        if sale_warehouse != wh_name_norm:
            continue
        art = item.get("supplierArticle", "—")
        price = float(item.get(price_type, 0))
        if art not in stat:
            stat[art] = {"qty": 0, "sum": 0.0, "price": price}
        stat[art]["qty"] += 1
        stat[art]["sum"] += price

    arts = list(stat.items())
    total_pages = max(1, (len(arts) + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    end = start + page_size
    arts_page = arts[start:end]

    # --- Форматирование Telegram-отчёта ---
    text = (
        f"🏭 <b>Склад: {wh_name}</b>\n"
        f"💶 <b>Цена:</b> {price_type_name}\n"
        f"🗓 <b>Период:</b> {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}\n"
        f"📄 <b>Страница:</b> {page}/{total_pages}\n\n"
    )
    total_qty = 0
    total_sum = 0.0
    if arts_page:
        for art, st in arts_page:
            text += (
                f"📦 <b>{art}</b>\n"
                f"• Кол-во: <b>{st['qty']}</b>\n"
                f"• Цена: <b>{st['price']:,.2f}</b>\n"
                f"• Сумма: <b>{st['sum']:,.2f}</b>\n\n"
            )
            total_qty += st['qty']
            total_sum += st['sum']
        sum_text = f"{total_sum:,.2f}".replace(",", " ")
        # Итог по странице
        text += f"<b>Итого на странице: {total_qty} шт / {sum_text} ₽</b>\n"
    else:
        text += "<i>Нет продаж по выбранным артикулам.</i>\n"

    # Итог по ВСЕМ артикулам этого склада — только на последней странице
    if page == total_pages:
        all_qty = sum(st['qty'] for st in stat.values())
        all_sum = sum(st['sum'] for st in stat.values())
        all_sum_txt = f"{all_sum:,.2f}".replace(",", " ")
        text += f"\n<b>🧾 ИТОГО за период по складу: {all_qty} шт / {all_sum_txt} ₽</b>"

    kb = build_pagination_keyboard(len(arts), page, page_size, "sales_wh_page:", f"sales_wh_menu:{warehouse_id}")
    kb.inline_keyboard.append([
        InlineKeyboardButton(text="📊 Экспорт в Excel", callback_data="sales_wh_export_xlsx")
    ])

    await callback.message.answer(text, reply_markup=kb, parse_mode="HTML")


    await progress_message.answer(text, reply_markup=kb, parse_mode="HTML")
@router.callback_query(F.data == "sales_wh_export_xlsx")
async def sales_wh_export_xlsx(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    sales = data.get("sales_report", [])
    warehouse_id = data.get("warehouse_id")
    date_from = data.get("date_from")
    date_to = data.get("date_to")
    warehouses = await get_cached_warehouses_dicts()
    wh = next((w for w in warehouses if str(w["id"]) == str(warehouse_id)), None)
    wh_name = wh.get("name", f"ID {warehouse_id}") if wh else f"ID {warehouse_id}"
    wh_name_norm = normalize_warehouse_name(wh_name)

    user_id = callback.from_user.id
    price_type = await get_user_price_type(user_id)
    price_type_name = price_type_human(price_type)

    # Собираем по артикулам: qty, price, sum
    stat = {}
    for item in sales:
        sale_warehouse = normalize_warehouse_name(item.get("warehouseName", ""))
        if sale_warehouse != wh_name_norm:
            continue
        art = item.get("supplierArticle", "—")
        price_val = float(item.get(price_type, 0))
        if art not in stat:
            stat[art] = {"qty": 0, "sum": 0.0, "price": price_val}
        stat[art]["qty"] += 1
        stat[art]["sum"] += price_val

    # Создаем Excel-файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Продажи по складу"

    # Шапка с видом цены и периодом
    ws.append([f"Цена: {price_type_name}"])
    ws.append([f"Период отчёта: {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"])
    ws.append([])

    # Заголовки
    ws.append(["Склад", "Артикул", "Кол-во", "Цена", "Сумма"])
    for cell in ws[4]:
        cell.font = Font(bold=True)

    total_qty = 0
    total_sum = 0.0
    if stat:
        for art, d in stat.items():
            ws.append([wh_name, art, d["qty"], d["price"], d["sum"]])
            total_qty += d["qty"]
            total_sum += d["sum"]
        ws.append([f"Итого по складу: {wh_name}", "", total_qty, "", total_sum])
    else:
        ws.append([wh_name, "Нет продаж", "", "", ""])

    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file = BufferedInputFile(buffer.read(), filename="sales_warehouse_report.xlsx")
    await callback.message.answer_document(file)

from storage.users import get_user_price_type
from bot.keyboards.keyboards import price_type_human

async def show_sales_report_all_warehouses(
        callback,
        date_from,
        date_to,
        state: FSMContext = None,
        page: int = 1
):
    user_id = callback.from_user.id
    price_type = await get_user_price_type(user_id)
    price_type_name = price_type_human(price_type)

    # Заголовок периода и цены
    if date_from == date_to:
        period_text = f"<b>{date_from.strftime('%d.%m.%Y')}</b>"
    else:
        period_text = f"<b>{date_from.strftime('%d.%m.%Y')}</b> — <b>{date_to.strftime('%d.%m.%Y')}</b>"

    progress_message = await callback.message.edit_text(
        f"⏳ Формируем отчёт по всем складам за период {period_text}.\n"
        f"💶 <b>Цена:</b> {price_type_name}\n"
        f"Пожалуйста, дождитесь завершения выгрузки!",
        parse_mode="HTML"
    )

    data = await state.get_data()
    need_api = (
            'all_sales_report' not in data or
            'all_warehouses' not in data or
            'date_from' not in data or
            'date_to' not in data or
            data.get("date_from") != date_from or
            data.get("date_to") != date_to
    )

    if need_api:
        warehouses = await get_cached_warehouses_dicts()
        api_key = await get_user_api_key(callback.from_user.id)
        while True:
            result = await get_sales_report_with_eta(api_key, date_from, date_to)
            if isinstance(result, dict) and result.get("error") == "ratelimit":
                retry = result["retry"]
                await progress_message.edit_text(
                    f"⏳ Формируем отчёт по всем складам за период {period_text}.\n"
                    f"💶 <b>Цена:</b> {price_type_name}\n"
                    f"   Отчет будет загружен не ранее чем через: <b>{retry} сек.</b> ⏳",
                    parse_mode="HTML"
                )
                await asyncio.sleep(retry)
                continue
            elif isinstance(result, dict) and result.get("error"):
                await progress_message.edit_text("❌ Ошибка при получении данных по всем складам.", parse_mode="HTML")
                return
            else:
                sales = result
                break

        # Фильтруем склады с продажами
        filtered_warehouses = []
        for wh in warehouses:
            wh_name_norm = normalize_warehouse_name(wh["name"])
            if any(normalize_warehouse_name(item.get("warehouseName", "")) == wh_name_norm for item in sales):
                filtered_warehouses.append(wh)

        await state.update_data(
            all_sales_report=sales,
            all_warehouses=filtered_warehouses,
            date_from=date_from,
            date_to=date_to
        )
        data = await state.get_data()

    # --- Используем данные из FSM ---
    sales = data.get("all_sales_report", [])
    filtered_warehouses = data.get("all_warehouses", [])
    date_from = data.get("date_from")
    date_to = data.get("date_to")

    PAGE_SIZE = 5
    total_wh = len(filtered_warehouses)
    total_pages = max(1, (total_wh + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(1, min(page, total_pages))
    start = (page - 1) * PAGE_SIZE
    end = start + PAGE_SIZE
    wh_page = filtered_warehouses[start:end]

    # --- Итоги по всем складам ---
    global_qty = 0
    global_sum = 0.0
    for wh in filtered_warehouses:
        wh_name = wh["name"]
        stat = {}
        wh_name_norm = normalize_warehouse_name(wh_name)
        for item in sales:
            sale_warehouse = normalize_warehouse_name(item.get("warehouseName", ""))
            if sale_warehouse != wh_name_norm:
                continue
            art = item.get("supplierArticle", "—")
            price_val = float(item.get(price_type, 0))
            if art not in stat:
                stat[art] = {"qty": 0, "sum": 0.0, "price": price_val}
            stat[art]["qty"] += 1
            stat[art]["sum"] += price_val
        global_qty += sum(v["qty"] for v in stat.values())
        global_sum += sum(v["sum"] for v in stat.values())

    # --- Форматирование Telegram-отчёта ---
    text = (
        f"<b>📦 Продажи по всем складам</b>\n"
        f"💶 <b>Цена:</b> {price_type_name}\n"
        f"🗓 <b>Период:</b> {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}\n"
        f"📄 <b>Страница:</b> {page}/{total_pages}\n\n"
    )

    for wh in wh_page:
        wh_name = wh["name"]
        stat = {}
        wh_name_norm = normalize_warehouse_name(wh_name)
        for item in sales:
            sale_warehouse = normalize_warehouse_name(item.get("warehouseName", ""))
            if sale_warehouse != wh_name_norm:
                continue
            art = item.get("supplierArticle", "—")
            price_val = float(item.get(price_type, 0))
            if art not in stat:
                stat[art] = {"qty": 0, "sum": 0.0, "price": price_val}
            stat[art]["qty"] += 1
            stat[art]["sum"] += price_val
        total_qty = sum(d["qty"] for d in stat.values())
        total_sum = sum(d["sum"] for d in stat.values())
        text += f"🏬 <b>{wh_name}</b>\n"
        if stat:
            for art, d in stat.items():
                text += (
                    f"📦 <b>{art}</b>\n"
                    f"• Кол-во: <b>{d['qty']}</b>\n"
                    f"• Цена: <b>{d['price']:,.2f}</b>\n"
                    f"• Сумма: <b>{d['sum']:,.2f}</b>\n"
                )
            sum_txt = f"{total_sum:,.2f}".replace(",", " ")
            text += f"\n<b>Итого по складу: {total_qty} шт / {sum_txt} ₽</b>\n\n"
        else:
            text += "   <i>Нет продаж</i>\n\n"

    # Итог по всем только на последней странице
    text += f"<i>Страница {page}/{total_pages}</i>"
    if page == total_pages:
        sum_text = f"{global_sum:,.2f}".replace(",", " ")
        text += (
            f"\n<b>🧾 ИТОГО за период по всем складам:</b> "
            f"{global_qty} шт / {sum_text} ₽"
        )

    kb = build_pagination_keyboard(total_wh, page, PAGE_SIZE, "sales_all_wh_page:", "main_sales_by_warehouses")
    kb.inline_keyboard.append([InlineKeyboardButton(text="📥 Экспорт в Excel", callback_data="sales_all_wh_export_xlsx")])
    await progress_message.edit_text(text, reply_markup=kb, parse_mode="HTML")

@router.callback_query(F.data.startswith("sales_all_wh_page:"))
async def all_wh_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split(":")[1])
    data = await state.get_data()
    date_from = data.get("date_from")
    date_to = data.get("date_to")
    await show_sales_report_all_warehouses(callback, date_from, date_to, state=state, page=page)

@router.callback_query(F.data == "sales_all_wh_export_xlsx")
async def sales_all_wh_export_xlsx(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    sales = data.get("all_sales_report", [])
    filtered_warehouses = data.get("all_warehouses", [])
    date_from = data.get("date_from")
    date_to = data.get("date_to")

    user_id = callback.from_user.id
    price_type = await get_user_price_type(user_id)
    price_type_name = price_type_human(price_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Продажи по складам"

    # Первая строка — тип цены и период жирным шрифтом
    ws.append([f"Цена: {price_type_name}"])
    ws.append([f"Период отчёта: {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"])
    ws.append([])

    # Вторая строка — заголовки (сумма и цена)
    ws.append(["Склад", "Артикул", "Кол-во продаж", "Цена", "Сумма"])
    for cell in ws[4]:
        cell.font = Font(bold=True)

    row = 5
    for wh in filtered_warehouses:
        wh_name = wh["name"]
        stat = {}
        wh_name_norm = normalize_warehouse_name(wh_name)
        for item in sales:
            sale_warehouse = normalize_warehouse_name(item.get("warehouseName", ""))
            if sale_warehouse != wh_name_norm:
                continue
            art = item.get("supplierArticle", "—")
            price_val = float(item.get(price_type, 0))
            if art not in stat:
                stat[art] = {"qty": 0, "sum": 0.0, "price": price_val}
            stat[art]["qty"] += 1
            stat[art]["sum"] += price_val
        total_qty = sum(d["qty"] for d in stat.values())
        total_sum = sum(d["sum"] for d in stat.values())
        if stat:
            for art, d in stat.items():
                ws.append([wh_name, art, d["qty"], d["price"], d["sum"]])
                row += 1
            ws.append([f"Итого по складу: {wh_name}", "", total_qty, "", total_sum])
            row += 1
        else:
            ws.append([wh_name, "Нет продаж", "", "", ""])
            row += 1
        ws.append([])  # Пустая строка между складами
        row += 1

    # Автоподбор ширины столбцов по содержимому
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Сохраняем файл в память
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file = BufferedInputFile(buffer.read(), filename="sales_all_warehouses_report.xlsx")
    await callback.message.answer_document(file)
